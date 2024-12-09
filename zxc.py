from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import os
from datetime import datetime
import time

# 오늘 날짜 가져오기
today_date = datetime.now().strftime('%Y-%m-%d')

# 엑셀 파일 경로 설정
result_folder = 'result'
excel_filename = f"가구사_일일매출_{today_date}.xlsx"
excel_path = os.path.join(result_folder, excel_filename)

# Selenium 크롬 드라이버 설정
driver_path = './chromedriver'  # 크롬 드라이버는 같은 폴더에 있어야 함
driver = webdriver.Chrome(executable_path=driver_path)

try:
    # 네이버 스마트스토어 페이지로 이동
    driver.get('https://sell.smartstore.naver.com/#/bizadvisor/summary/daily')

    # 페이지 로딩 대기
    time.sleep(5)

    # 데이터 추출 (예: 일일 매출 데이터를 특정 CSS 셀렉터에서 가져오기)
    # 셀렉터는 페이지 HTML 구조에 맞게 수정 필요
    daily_sales_element = driver.find_element(By.CSS_SELECTOR, 'selector_for_daily_sales')
    daily_sales_data = daily_sales_element.text.strip()  # 데이터를 문자열로 가져옴

    print(f"추출된 데이터: {daily_sales_data}")

    # 엑셀 파일에 데이터 기록
    if not os.path.exists(excel_path):
        print(f"파일 {excel_path}이(가) 존재하지 않습니다.")
    else:
        workbook = load_workbook(excel_path)
        worksheet = workbook['평일']  # '평일' 시트 선택
        ay_column = 'AY'

        # 데이터를 AY 열의 마지막 빈 행에 추가
        row_to_write = worksheet.max_row + 1
        worksheet[f"{ay_column}{row_to_write}"] = daily_sales_data

        # 파일 저장
        workbook.save(excel_path)
        print(f"데이터가 {excel_path} 파일의 '평일' 시트의 AY 열에 기록되었습니다.")

finally:
    # 드라이버 종료
    driver.quit()
