import os
import sys
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
import time
from datetime import datetime
import re

def get_current_directory():
    """현재 실행 파일 또는 스크립트의 디렉터리 경로 반환"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def setup_directories():
    """경로 설정 및 결과 저장 폴더 생성"""
    current_dir = get_current_directory()
    chrome_driver_path = os.path.join(current_dir, 'chromedriver.exe')
    excel_file_path = os.path.join(current_dir, 'data.xlsx')
    template_file_path = os.path.join(current_dir, '가구사_일일매출_템플릿.xlsx')
    output_dir = os.path.join(current_dir, 'result')
    os.makedirs(output_dir, exist_ok=True)
    return chrome_driver_path, excel_file_path, template_file_path, output_dir

def initialize_webdriver(chrome_driver_path):
    """웹 드라이버 초기화"""
    service = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=service)
    wait = WebDriverWait(driver, 10)
    return driver, wait

def read_excel_data(excel_file_path):
    """엑셀 파일에서 데이터 읽기"""
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb['계정정보']
    links_and_filenames = [
        (sheet[f'B{row}'].value, sheet[f'A{row}'].value, sheet[f'C{row}'].value, sheet[f'D{row}'].value)
        for row in range(2, sheet.max_row + 1)
        if sheet[f'B{row}'].value and sheet[f'A{row}'].value
    ]
    return links_and_filenames

def login_to_website(driver, wait, link, user_id, password):
    """웹사이트 로그인"""
    driver.get(link)
    try:
        user_id_input = wait.until(EC.presence_of_element_located((By.NAME, 'master_id')))
        password_input = wait.until(EC.presence_of_element_located((By.NAME, 'master_pw')))
        user_id_input.send_keys(user_id)
        password_input.send_keys(password)
        password_input.send_keys(Keys.RETURN)
        time.sleep(3)  # 페이지 로드 대기
    except Exception as e:
        print(f"로그인 중 오류 발생: {e}")
        raise

def extract_data_from_page(driver):
    """웹 페이지 데이터 추출"""
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, 'html.parser')

    # 주문정보 추출
    order_data = []
    orders = soup.select('#today_order tr')
    for row in orders[1:]:  # 첫 번째 행은 헤더
        cols = [col.text.strip() for col in row.find_all('td')]
        order_data.append(cols)

    # 에누리신청리스트 추출
    discount_data = []
    discounts = soup.select('#today_discount tr')
    for row in discounts[1:]:
        cols = [col.text.strip() for col in row.find_all('td')]
        discount_data.append(cols)

    # 견적리스트 추출
    estimate_data = []
    estimates = soup.select('#today_online tr')
    for row in estimates[1:]:
        cols = [col.text.strip() for col in row.find_all('td')]
        estimate_data.append(cols)

    # 신규 데이터 추출
    new_inquiries = soup.select_one("#mentomen_new_div").get_text(separator="\n").strip() if soup.select_one("#mentomen_new_div") else "데이터 없음"
    new_product_inquiries = soup.select_one("#pfo_new_div").get_text(separator="\n").strip() if soup.select_one("#pfo_new_div") else "데이터 없음"
    new_visit_requests = soup.select_one("#ele_new_div").get_text(separator="\n").strip() if soup.select_one("#ele_new_div") else "데이터 없음"
    new_layout_requests = soup.select_one("#lay_new_div").get_text(separator="\n").strip() if soup.select_one("#lay_new_div") else "데이터 없음"

    return {
        "order_data": order_data,
        "discount_data": discount_data,
        "estimate_data": estimate_data,
        "new_inquiries": new_inquiries,
        "new_product_inquiries": new_product_inquiries,
        "new_visit_requests": new_visit_requests,
        "new_layout_requests": new_layout_requests
    }

def extract_name(full_string):
    """문자열에서 []를 제외한 이름 부분만 추출"""
    if "[" in full_string:
        return full_string.split("[")[0].strip()
    return full_string.strip()

def clean_price(price_str):
    """가격 문자열에서 '\'과 '원' 제거 및 숫자만 반환"""
    return re.sub(r'[^0-9]', '', price_str)

def save_to_individual_excel(extracted_data, filename, output_dir):
    """각 업체명에 해당하는 데이터를 별도의 엑셀 파일로 저장"""
    output_file_path = os.path.join(output_dir, f"{filename}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # 노란색 셀 스타일
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 신규 데이터를 추가
    ws.append(["신규 1:1 문의", extracted_data["new_inquiries"]])
    ws.append(["신규 상품 문의", extracted_data["new_product_inquiries"]])
    ws.append(["신규 방문견적 신청", extracted_data["new_visit_requests"]])
    ws.append(["신규 레이아웃 신청", extracted_data["new_layout_requests"]])
    ws.append([])

    # 주문정보 추가
    order_start_row = None
    order_names = []
    if extracted_data["order_data"]:
        ws.append(["주문정보"])
        ws.append(["주문정보", "주문접수일", "배송방법", "주문자/입금자", "회원구분", "결제금액", "진행상태", "결제종류"])
        order_start_row = ws.max_row + 1
        for row in extracted_data["order_data"]:
            ws.append(row)
            order_names.append(extract_name(row[0]))  # 주문자 이름 추출
        ws.append([])

    # 에누리신청리스트 추가 및 중복 제거
    if extracted_data["discount_data"]:
        ws.append(["에누리신청리스트"])
        ws.append(["의뢰인", "요청금액", "연락처", "최초상담원"])
        for row in extracted_data["discount_data"]:
            name = extract_name(row[0])
            if name in order_names:  # 중복된 경우 스킵
                continue
            ws.append(row)
            # 중복 항목은 노란색으로 처리
            if name in order_names:
                for col in range(1, len(row) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = yellow_fill
        ws.append([])

    # 견적리스트 추가 및 중복 제거
    if extracted_data["estimate_data"]:
        ws.append(["견적리스트"])
        ws.append(["의뢰인", "총금액", "연락처", "최초상담원"])
        for row in extracted_data["estimate_data"]:
            name = extract_name(row[0])
            if name in order_names:  # 중복된 경우 스킵
                continue
            ws.append(row)
            # 중복 항목은 노란색으로 처리
            if name in order_names:
                for col in range(1, len(row) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = yellow_fill
        ws.append([])

    # 엑셀 파일 저장
    wb.save(output_file_path)
    print(f"{output_file_path}에 데이터가 저장되었습니다.")

    """각 업체명에 해당하는 데이터를 별도의 엑셀 파일로 저장"""
    output_file_path = os.path.join(output_dir, f"{filename}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # 노란색 셀 스타일
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 신규 데이터를 추가
    ws.append(["신규 1:1 문의", extracted_data["new_inquiries"]])
    ws.append(["신규 상품 문의", extracted_data["new_product_inquiries"]])
    ws.append(["신규 방문견적 신청", extracted_data["new_visit_requests"]])
    ws.append(["신규 레이아웃 신청", extracted_data["new_layout_requests"]])
    ws.append([])

    # 주문자 이름 목록 추출
    order_names = [extract_name(row[0]) for row in extracted_data["order_data"]]

    # 주문정보 추가
    if extracted_data["order_data"]:
        ws.append(["주문정보"])
        ws.append(["주문정보", "주문접수일", "배송방법", "주문자/입금자", "회원구분", "결제금액", "진행상태", "결제종류"])
        for row in extracted_data["order_data"]:
            ws.append(row)
        ws.append([])

    # 에누리신청리스트 추가
    if extracted_data["discount_data"]:
        ws.append(["에누리신청리스트"])
        ws.append(["의뢰인", "요청금액", "연락처", "최초상담원"])
        for row in extracted_data["discount_data"]:
            name = extract_name(row[0])
            ws.append(row)
            if name in order_names:
                for col in range(1, len(row) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = yellow_fill
        ws.append([])

    # 견적리스트 추가
    if extracted_data["estimate_data"]:
        ws.append(["견적리스트"])
        ws.append(["의뢰인", "총금액", "연락처", "최초상담원"])
        for row in extracted_data["estimate_data"]:
            name = extract_name(row[0])
            ws.append(row)
            if name in order_names:
                for col in range(1, len(row) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = yellow_fill
        ws.append([])

    # 엑셀 파일 저장
    wb.save(output_file_path)
    print(f"{output_file_path}에 데이터가 저장되었습니다.")

def save_to_template_excel(output_dir, template_file_path):
    """템플릿 기반으로 데이터를 일일 매출 파일에 저장"""
    today_str = datetime.now().strftime("%Y%m%d")
    output_file_path = os.path.join(output_dir, f"가구사_일일매출_{today_str}.xlsx")

    # 템플릿 파일을 불러와서 새로운 워크북으로 사용
    wb = openpyxl.load_workbook(template_file_path)
    ws = wb["평일"]

    # 업체별 파일에서 데이터 읽어오기
    for root, _, files in os.walk(output_dir):
        for file in files:
            if file.endswith(".xlsx") and file != f"가구사_일일매출_{today_str}.xlsx":
                individual_file_path = os.path.join(root, file)
                individual_wb = openpyxl.load_workbook(individual_file_path)
                individual_ws = individual_wb.active

                # 업체명 가져오기 (파일명에서 추출)
                company_name = file.replace(".xlsx", "")

                # 매칭되는 업체명 찾기
                for row in range(1, ws.max_row + 1):
                    if ws[f"A{row}"].value == company_name:
                        # 주문정보 결제금액 추가 (특정 값 '결제금액' 찾기)
                        payment_row = None
                        for r in individual_ws.iter_rows(min_row=1, max_row=individual_ws.max_row, min_col=1, max_col=individual_ws.max_column):
                            if any(cell.value == "결제금액" for cell in r):
                                payment_row = r[0].row
                                break

                        if payment_row:
                            current_column = 18  # R열에 해당
                            order_data = []
                            for i in range(payment_row + 1, individual_ws.max_row + 1):
                                value = individual_ws.cell(row=i, column=6).value
                                if value is None or value == '':
                                    break
                                order_data.append(clean_price(value))

                            for value in order_data:
                                try:
                                    numeric_value = int(value)
                                except ValueError:
                                    numeric_value = 0
                                ws.cell(row=row, column=current_column, value=numeric_value)
                                current_column += 1

                        # 에누리 요청금액 추가 (특정 값 '요청금액' 찾기)
                        request_row = None
                        for r in individual_ws.iter_rows(min_row=1, max_row=individual_ws.max_row, min_col=1, max_col=individual_ws.max_column):
                            if any(cell.value == "요청금액" for cell in r):
                                request_row = r[0].row
                                break

                        if request_row:
                            current_column = 52  # AZ열에 해당
                            discount_data = []
                            for i in range(request_row + 1, individual_ws.max_row + 1):
                                value = individual_ws.cell(row=i, column=2).value
                                if value is None or value == '':
                                    break
                                discount_data.append(clean_price(value))

                            for value in discount_data:
                                try:
                                    numeric_value = int(value)
                                except ValueError:
                                    numeric_value = 0
                                ws.cell(row=row, column=current_column, value=numeric_value)
                                current_column += 1

                        # 견적 총금액 추가 (특정 값 '총금액' 찾기)
                        estimate_row = None
                        for r in individual_ws.iter_rows(min_row=1, max_row=individual_ws.max_row, min_col=1, max_col=individual_ws.max_column):
                            if any(cell.value == "총금액" for cell in r):
                                estimate_row = r[0].row
                                break

                        if estimate_row:
                            # 에누리 요청금액 뒤에 이어서 견적 총금액 추가
                            estimate_data = []
                            for i in range(estimate_row + 1, individual_ws.max_row + 1):
                                value = individual_ws.cell(row=i, column=2).value
                                if value is None or value == '':
                                    break
                                estimate_data.append(clean_price(value))

                            for value in estimate_data:
                                try:
                                    numeric_value = int(value)
                                except ValueError:
                                    numeric_value = 0
                                ws.cell(row=row, column=current_column, value=numeric_value)
                                current_column += 1

    # 엑셀 파일 저장
    wb.save(output_file_path)
    print(f"{output_file_path}에 템플릿 기반 데이터가 저장되었습니다.")





def main():
    # 디렉터리 및 경로 설정
    chrome_driver_path, excel_file_path, template_file_path, output_dir = setup_directories()

    # 웹 드라이버 초기화
    driver, wait = initialize_webdriver(chrome_driver_path)

    try:
        # 엑셀 데이터 읽기
        links_and_filenames = read_excel_data(excel_file_path)

        for link, filename, user_id, password in links_and_filenames:
            try:
                print(f"작업 중: {filename}")
                # 로그인
                login_to_website(driver, wait, link, user_id, password)

                # 데이터 추출
                extracted_data = extract_data_from_page(driver)

                # 업체명별 파일 저장
                save_to_individual_excel(extracted_data, filename, output_dir)

                # 템플릿 기반 파일 저장
                save_to_template_excel(output_dir, template_file_path)

            except Exception as e:
                print(f"오류 발생: {e}")
                continue

    finally:
        driver.quit()

if __name__ == "__main__":
    main()
